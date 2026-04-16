# ============================================================
# ENDPOINT REPORTE DE COMISIONES
# Archivo: app/bills/routes_reporte.py
#
# Uso:
#   GET /api/billing/reporte-comisiones
#       ?sede_id=SD-88809
#       &fecha_desde=01/05/2026          ← acepta DD/MM/YYYY o YYYY-MM-DD
#       &fecha_hasta=15/05/2026
#       &tipo_item=ambos                 ← servicios | productos | ambos
#
# Hojas generadas:
#   1. Detalle de transacciones — una fila por ítem
#
# Comisiones de productos (por niveles, calculadas sobre el total
# de unidades vendidas por persona en el período consultado):
#   Nivel 1  (1 – 5  uds): 2 %
#   Nivel 2  (6 – 10 uds): 3 %
#   Nivel 3 (11 – 20 uds): 4 %
#   Nivel 4  (> 20   uds): 5 %
#
# Registro en main.py:
#   from app.bills.routes_reporte import router as reporte_router
#   app.include_router(reporte_router, prefix="/api/billing")
# ============================================================

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.auth.routes import get_current_user
from app.database.mongo import (
    collection_sales,
    collection_citas,
    collection_locales,
    collection_estilista,
    collection_auth,
    collection_card
)
from bson import ObjectId
import boto3
import os
from urllib.parse import urlparse
from app.scheduling.submodules.quotes.controllers import (
    generar_pdf_ficha,
    crear_html_correo_ficha,
    enviar_correo_con_pdf,
    descargar_imagen,           # ← agregar
    comprimir_imagen_para_pdf,  # ← agregar
)

router = APIRouter()
 
# ── Paleta ────────────────────────────────────────────────────────
_NEGRO  = "1A1A1A"
_BLANCO = "FFFFFF"
_AZUL   = "0D2137"
_AZUL_M = "1A3A5C"
_VERDE  = "0A4D2E"
_VERDE_M = "1E7145"
_VERDE_L = "E6F4ED"
_GRIS_B = "FAFAFA"
_AMBAR  = "FFF8E6"
 
MNY = "#,##0"
PCT = "0.0%"
INT = "#,##0"
DAT = "DD/MM/YYYY"
 
 
# ── Helpers de estilo ─────────────────────────────────────────────
 
def _bord():
    s = Side(style="thin", color="D0D0D0")
    return Border(top=s, bottom=s, left=s, right=s)
 
 
def _cell(ws, row, col, val, bg=_BLANCO, fg=_NEGRO, bold=False,
          fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=9, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _bord()
    if fmt:
        c.number_format = fmt
    return c
 
 
def _banner(ws, row, text, bg, ncols, height=36):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", size=14, bold=True, color=_BLANCO)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height
 
 
def _meta(ws, row, text, ncols, color="555555", height=18):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", size=10, color=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height
 
 
def _hdr(ws, row, col, text, bg):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Arial", size=9, bold=True, color=_BLANCO)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _bord()
    return c
 
 
# ── Helpers de lógica ─────────────────────────────────────────────
 
def _parse_fecha(valor: str) -> datetime:
    """Acepta YYYY-MM-DD o DD-MM-YYYY."""
    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(valor.strip(), fmt)
        except ValueError:
            continue
    raise ValueError(
        f"Formato de fecha no reconocido: '{valor}'. "
        "Use YYYY-MM-DD o DD-MM-YYYY."
    )
 
 
def _tier_rate(total_unidades: int) -> float:
    """Devuelve la tasa de comisión según el nivel de ventas de productos."""
    if total_unidades <= 5:
        return 0.02   # Nivel 1
    elif total_unidades <= 10:
        return 0.03   # Nivel 2
    elif total_unidades <= 20:
        return 0.04   # Nivel 3
    else:
        return 0.05   # Nivel 4 — vendedor estrella
 
 
def _nivel_label(total_unidades: int) -> str:
    if total_unidades <= 5:
        return "Nv.1 (2%)"
    elif total_unidades <= 10:
        return "Nv.2 (3%)"
    elif total_unidades <= 20:
        return "Nv.3 (4%)"
    else:
        return "Nv.4 (5%)"
 
 
# ── Endpoint ──────────────────────────────────────────────────────
 
@router.get("/reporte-comisiones")
async def reporte_comisiones(
    sede_id:      str           = Query(...,    description="Ej: SD-88809"),
    fecha_desde:  Optional[str] = Query(None,   description="YYYY-MM-DD o DD/MM/YYYY"),
    fecha_hasta:  Optional[str] = Query(None,   description="YYYY-MM-DD o DD/MM/YYYY"),
    tipo_item:    str           = Query("ambos", description="servicios | productos | ambos"),
    current_user: dict          = Depends(get_current_user),
):
    if current_user["rol"] not in ["admin_sede", "super_admin"]:
        raise HTTPException(status_code=403, detail="No autorizado")
 
    # ── Validar tipo_item ─────────────────────────────────────────
    tipo_item = tipo_item.lower().strip()
    if tipo_item not in ("servicios", "productos", "ambos"):
        raise HTTPException(
            status_code=422,
            detail="tipo_item debe ser 'servicios', 'productos' o 'ambos'."
        )
 
    # ── Parsear fechas (acepta DD/MM/YYYY o YYYY-MM-DD) ───────────
    hoy = datetime.now()
    try:
        dt_desde = _parse_fecha(fecha_desde) if fecha_desde else hoy.replace(day=1)
        dt_hasta = (
            _parse_fecha(fecha_hasta).replace(hour=23, minute=59, second=59)
            if fecha_hasta
            else hoy.replace(hour=23, minute=59, second=59)
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
 
    desde_display = dt_desde.strftime("%d/%m/%Y")
    hasta_display = dt_hasta.strftime("%d/%m/%Y")
    desde_str     = dt_desde.strftime("%Y-%m-%d")
    hasta_str     = dt_hasta.strftime("%Y-%m-%d")
 
    # ── Sede ──────────────────────────────────────────────────────
    sede = await collection_locales.find_one({"sede_id": sede_id})
    if not sede:
        raise HTTPException(status_code=404, detail="Sede no encontrada")
    sede_nombre = sede.get("nombre", sede_id)
 
    # ── Mapa email → nombre real ──────────────────────────────────
    auth_docs = await collection_auth.find(
        {"sede_id": sede_id},
        {"correo_electronico": 1, "email": 1, "nombre": 1},
    ).to_list(None)
 
    email_a_nombre: dict = {}
    for a in auth_docs:
        email = str(a.get("correo_electronico") or a.get("email") or "").strip().lower()
        nombre_auth = str(a.get("nombre") or "").strip()
        if email and nombre_auth:
            email_a_nombre[email] = nombre_auth
 
    def resolver_email(valor: str) -> str:
        v = str(valor or "").strip()
        if not v:
            return v
        if "@" in v:
            return email_a_nombre.get(v.lower(), v.split("@")[0])
        return v
 
    # ── Mapa nombre ↔ profesional_id ─────────────────────────────
    est_docs = await collection_estilista.find({"sede_id": sede_id}).to_list(None)
    pid_a_nombre: dict = {}
    nombre_a_pid: dict = {}
    for e in est_docs:
        pid    = e.get("profesional_id", "")
        nombre = (e.get("nombre", "") + " " + e.get("apellido", "")).strip()
        if pid and nombre:
            pid_a_nombre[pid]            = nombre
            nombre_a_pid[nombre.lower()] = pid
 
    # ── Obtener ventas del período ────────────────────────────────
    # Citas: se filtra por fecha de la cita (agenda), no por fecha_pago
    citas_docs = await collection_citas.find(
        {
            "sede_id":        sede_id,
            "estado_factura": "facturado",
            "fecha":          {"$gte": desde_str, "$lte": hasta_str},
        },
        {"_id": 1, "fecha": 1},
    ).to_list(None)
 
    cita_id_a_fecha: dict = {}
    cita_ids_str = []
    for c in citas_docs:
        oid = str(c["_id"])
        cita_ids_str.append(oid)
        cita_id_a_fecha[oid] = c.get("fecha", "")
 
    ventas_citas = []
    if cita_ids_str:
        ventas_citas = await collection_sales.find(
            {
                "sede_id":     sede_id,
                "tipo_origen": "cita",
                "origen_id":   {"$in": cita_ids_str},
            }
        ).to_list(None)
 
    ventas_directas = await collection_sales.find(
        {
            "sede_id":        sede_id,
            "estado_factura": "facturado",
            "tipo_origen":    {"$ne": "cita"},
            "fecha_pago":     {"$gte": dt_desde, "$lte": dt_hasta},
        }
    ).to_list(None)
 
    ventas = ventas_citas + ventas_directas
 
    # ── Nombres que no son personas reales ───────────────────────
    nombres_sede = {sede_nombre.strip().lower(), sede_id.strip().lower()}
    ROLES_PROPIOS = {"recepcionista", "call_center", "admin_sede"}
 
    # ── Helper: resolver responsable de productos en una venta ────
    def _responsable_producto(v: dict, item: dict, responsable_srv: Optional[str]) -> Optional[str]:
        """
        Devuelve el nombre (display) del responsable del producto.
        Prioridad:
          1. agregado_por_rol en ROLES_PROPIOS → ese email resuelto
          2. Venta directa con vendido_por válido → vendido_por
          3. Cita → responsable del servicio
        """
        agr_rol   = str(item.get("agregado_por_rol", "") or "")
        agr_email = str(item.get("agregado_por_email", "") or "").strip()
        tipo_origen = v.get("tipo_origen", v.get("tipo_venta", ""))
 
        if agr_rol in ROLES_PROPIOS and agr_email:
            return resolver_email(agr_email)
 
        vendido_raw = str(v.get("vendido_por", "") or "").strip()
        vendido     = resolver_email(vendido_raw)
        local_venta = str(v.get("local", "") or "").strip().lower()
        if vendido.strip().lower() in nombres_sede or vendido.strip().lower() == local_venta:
            vendido = ""
 
        if tipo_origen != "cita" and vendido and "," not in vendido:
            return vendido
 
        return responsable_srv  # fallback: estilista de la cita
 
    # ══════════════════════════════════════════════════════════════
    # PASO 1: contar unidades de productos por persona (por nombre
    # normalizado) para determinar el nivel de comisión.
    # Se agrupa por nombre resuelto, sin importar si tiene o no
    # profesional_id, para cubrir todos los casos edge.
    # ══════════════════════════════════════════════════════════════
    prod_unidades: dict[str, int] = {}  # nombre_norm → total unidades
 
    for v in ventas:
        tipo_origen  = v.get("tipo_origen", v.get("tipo_venta", ""))
        pid_doc      = str(v.get("profesional_id", "") or "")
        vendido_raw  = str(v.get("vendido_por",   "") or "").strip()
        facturado_raw = str(v.get("facturado_por", "") or "").strip()
        local_venta  = str(v.get("local", "") or "").strip().lower()
 
        vendido  = resolver_email(vendido_raw)
        facturado = resolver_email(facturado_raw)
 
        if vendido.strip().lower() in nombres_sede or vendido.strip().lower() == local_venta:
            vendido = ""
 
        # Resolver responsable de servicio (para fallback de producto en cita)
        if pid_doc.startswith("ES-"):
            nom_srv = pid_a_nombre.get(pid_doc) or v.get("profesional_nombre", "") or pid_doc
        elif vendido and "," not in vendido:
            nom_srv = vendido
        elif facturado and "," not in facturado:
            nom_srv = facturado
        else:
            nom_srv = None
 
        for item in v.get("items", []):
            if item.get("tipo") != "producto":
                continue
 
            resp_prod = _responsable_producto(v, item, nom_srv)
            if not resp_prod:
                continue
 
            key_n = resp_prod.strip().lower()
            cant  = int(item.get("cantidad", 1))
            prod_unidades[key_n] = prod_unidades.get(key_n, 0) + cant
 
    # Mapa nombre_norm → tasa de comisión y etiqueta de nivel
    com_rate_map: dict[str, float] = {k: _tier_rate(v) for k, v in prod_unidades.items()}
    nivel_map:    dict[str, str]   = {k: _nivel_label(v) for k, v in prod_unidades.items()}
 
    # ══════════════════════════════════════════════════════════════
    # PASO 2: construir filas de detalle con comisiones de productos
    # recalculadas según el nivel alcanzado por cada persona.
    # ══════════════════════════════════════════════════════════════
    detalles = []
 
    TIPO_PERMITIDO = {
        "servicios": {"servicio"},
        "productos":  {"producto"},
        "ambos":      {"servicio", "producto"},
    }
    tipos_validos = TIPO_PERMITIDO[tipo_item]
 
    for v in ventas:
        tipo_origen   = v.get("tipo_origen", v.get("tipo_venta", ""))
        pid_v         = str(v.get("profesional_id", "") or "")
        pnombre_v     = v.get("profesional_nombre", "") or ""
        vendido_raw   = str(v.get("vendido_por",   "") or "").strip()
        facturado_raw = str(v.get("facturado_por", "") or "").strip()
        cliente       = v.get("nombre_cliente", "") or ""
        comp          = v.get("numero_comprobante", "") or ""
        fecha_v       = v.get("fecha_pago")
        local_venta   = str(v.get("local", "") or "").strip().lower()
 
        vendido  = resolver_email(vendido_raw)
        facturado = resolver_email(facturado_raw)
        if vendido.strip().lower() in nombres_sede or vendido.strip().lower() == local_venta:
            vendido = ""
 
        # Responsable principal (servicios / fallback productos en cita)
        if tipo_origen == "cita":
            responsable_srv = pnombre_v or pid_a_nombre.get(pid_v, pid_v) or ""
            tipo_label      = "Cita"
            origen_id_v     = str(v.get("origen_id", ""))
            fecha_cita_str  = cita_id_a_fecha.get(origen_id_v, "")
            if fecha_cita_str:
                try:
                    fecha_v = datetime.strptime(fecha_cita_str, "%Y-%m-%d")
                except ValueError:
                    pass
        else:
            if vendido and "," not in vendido:
                responsable_srv = vendido
            else:
                responsable_srv = resolver_email(facturado_raw)
            tipo_label = "Venta directa"
 
        for item in v.get("items", []):
            tipo_i = item.get("tipo", "")
            if tipo_i not in tipos_validos:
                continue
 
            sub  = float(item.get("subtotal", 0))
            cant = int(item.get("cantidad", 1))
 
            if tipo_i == "servicio":
                # Comisión de servicio: se respeta la almacenada en BD
                com_raw = float(item.get("comision", 0))
                pct_raw = float(item.get("porcentaje_comision",
                                item.get("comision_porcentaje", 0)))
                pct = pct_raw if pct_raw else (
                    round(com_raw / sub * 100, 2) if sub > 0 and com_raw > 0 else 0
                )
                com = com_raw
                responsable = responsable_srv
                nivel_txt   = ""
 
            else:  # producto
                responsable = _responsable_producto(v, item, responsable_srv) or ""
                key_n       = responsable.strip().lower()
                rate        = com_rate_map.get(key_n, 0.02)
                com         = round(sub * rate, 2)
                pct         = rate * 100
                nivel_txt   = nivel_map.get(key_n, "Nv.1 (2%)")
 
            detalles.append({
                "fecha":       fecha_v,
                "comprobante": comp,
                "tipo":        tipo_label,
                "responsable": responsable,
                "tipo_item":   "Servicio" if tipo_i == "servicio" else "Producto",
                "nivel_com":   nivel_txt,          # vacío para servicios
                "item_nom":    item.get("nombre", ""),
                "cant":        cant,
                "precio":      float(item.get("precio_unitario", 0)),
                "subtotal":    sub,
                "pct":         pct / 100,
                "comision":    com,
                "cliente":     cliente,
            })
 
    # ════════════════════════════════════════════════════════════════
    # EXCEL — una sola hoja: Detalle de transacciones
    # ════════════════════════════════════════════════════════════════
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle de transacciones"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    N = 13  # columnas
 
    # Subtítulo de filtro aplicado
    filtro_label = {
        "servicios": "Solo servicios",
        "productos":  "Solo productos",
        "ambos":      "Servicios y productos",
    }[tipo_item]
 
    _banner(ws, 1, "RIZOS FELICES — DETALLE DE TRANSACCIONES", _VERDE, N)
    _meta(ws, 2, f"Sede: {sede_nombre}", N, color=_VERDE_M)
    _meta(
        ws, 3,
        f"Período: {desde_display}  →  {hasta_display}   ·   "
        f"{filtro_label}   ·   "
        f"{len(detalles)} ítems   ·   "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        N, color="666666", height=16,
    )
    ws.row_dimensions[4].height = 38
 
    # Encabezados de columna
    h_cols = [
        ("Fecha",                11),
        ("Comprobante",          14),
        ("Tipo",                 13),
        ("Estilista / Vendedor", 22),
        ("Tipo ítem",            11),
        ("Nivel comisión",       13),   # solo aplica a productos
        ("Nombre ítem",          28),
        ("Cant.",                 7),
        ("Precio unit. (COP)",   16),
        ("Subtotal (COP)",       15),
        ("% comisión",           11),
        ("Comisión (COP)",       15),
        ("Cliente",              22),
    ]
    for ci, (t, w) in enumerate(h_cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        _hdr(ws, 4, ci, t, _VERDE_M)
 
    TIPO_COLOR = {
        "Cita":          ("EFF6FF", "1E40AF"),
        "Venta directa": ("F0FDF4", "166534"),
    }
    ITEM_BG = {"Servicio": "FEFCE8", "Producto": _BLANCO}
 
    r = 5
    for i, det in enumerate(detalles):
        bg = _GRIS_B if i % 2 == 0 else _BLANCO
 
        _cell(ws, r, 1, det["fecha"],       bg=bg, fmt=DAT, align="center")
        _cell(ws, r, 2, det["comprobante"], bg=bg,          align="center")
 
        # Columna Tipo — color según cita / venta directa
        bg_t, fg_t = TIPO_COLOR.get(det["tipo"], (_BLANCO, _NEGRO))
        ct = ws.cell(row=r, column=3, value=det["tipo"])
        ct.font      = Font(name="Arial", size=9, bold=True, color=fg_t)
        ct.fill      = PatternFill("solid", fgColor=bg_t)
        ct.alignment = Alignment(horizontal="center", vertical="center")
        ct.border    = _bord()
 
        _cell(ws, r, 4, det["responsable"], bg=bg)
 
        # Columna Tipo ítem — color diferenciado
        bg_i = ITEM_BG.get(det["tipo_item"], _BLANCO)
        ci_c = ws.cell(row=r, column=5, value=det["tipo_item"])
        ci_c.font      = Font(name="Arial", size=9, color=_NEGRO)
        ci_c.fill      = PatternFill("solid", fgColor=bg_i)
        ci_c.alignment = Alignment(horizontal="center", vertical="center")
        ci_c.border    = _bord()
 
        # Columna Nivel comisión (solo productos; vacío para servicios)
        bg_nv = "EDE9FE" if det["nivel_com"] else bg   # lavanda suave para productos
        fg_nv = "5B21B6" if det["nivel_com"] else _NEGRO
        c_nv  = ws.cell(row=r, column=6, value=det["nivel_com"] or "—")
        c_nv.font      = Font(name="Arial", size=9, bold=bool(det["nivel_com"]), color=fg_nv)
        c_nv.fill      = PatternFill("solid", fgColor=bg_nv)
        c_nv.alignment = Alignment(horizontal="center", vertical="center")
        c_nv.border    = _bord()
 
        _cell(ws, r, 7,  det["item_nom"], bg=bg)
        _cell(ws, r, 8,  det["cant"],     bg=bg, fmt=INT, align="center")
        _cell(ws, r, 9,  det["precio"],   bg=bg, fmt=MNY, align="right")
        _cell(ws, r, 10, det["subtotal"], bg=bg, fmt=MNY, align="right")
 
        # % comisión — ámbar si es 0 pero hay comisión (dato raro en BD)
        bg_pct = _AMBAR if det["pct"] == 0 and det["comision"] > 0 else bg
        _cell(ws, r, 11, det["pct"],      bg=bg_pct, fmt=PCT, align="center")
 
        # Comisión — verde si > 0
        bg_com = _VERDE_L if det["comision"] > 0 else bg
        fg_com = _VERDE_M if det["comision"] > 0 else _NEGRO
        _cell(ws, r, 12, det["comision"], bg=bg_com, fg=fg_com,
              bold=det["comision"] > 0, fmt=MNY, align="right")
 
        _cell(ws, r, 13, det["cliente"], bg=bg)
        r += 1
 
    # Fila de totales al final
    ws.row_dimensions[r].height = 20
    for ci in range(1, N + 1):
        c = ws.cell(row=r, column=ci)
        c.fill   = PatternFill("solid", fgColor="E8F5E9")
        c.font   = Font(name="Arial", size=9, bold=True, color=_VERDE_M)
        c.border = _bord()
    ws.cell(row=r, column=1, value="TOTALES").alignment = Alignment(
        horizontal="left", vertical="center"
    )
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="E8F5E9")
    ws.cell(row=r, column=1).font = Font(name="Arial", size=9, bold=True, color=_VERDE_M)
    for ci, cl in [(10, "J"), (12, "L")]:
        ws.cell(row=r, column=ci, value=f"=SUM({cl}5:{cl}{r - 1})").number_format = MNY
        ws.cell(row=r, column=ci).fill      = PatternFill("solid", fgColor="E8F5E9")
        ws.cell(row=r, column=ci).font      = Font(name="Arial", size=9, bold=True, color=_VERDE_M)
        ws.cell(row=r, column=ci).border    = _bord()
        ws.cell(row=r, column=ci).alignment = Alignment(horizontal="right", vertical="center")
 
    # ── Nota al pie ───────────────────────────────────────────────
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
    nota = (
        "NOTA: Comisiones de productos calculadas por niveles según unidades vendidas en el período "
        "(Nv.1 ≤5 uds → 2 % · Nv.2 6-10 → 3 % · Nv.3 11-20 → 4 % · Nv.4 >20 → 5 %). "
        "Comisiones de servicios respetan el porcentaje registrado en la venta."
    )
    ws.cell(row=r, column=1, value=nota).font = Font(
        name="Arial", size=8, italic=True, color="888888"
    )
    ws.cell(row=r, column=1).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    ws.row_dimensions[r].height = 32
 
    # ── Respuesta ─────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
 
    nombre_archivo = (
        f"comisiones_{sede_nombre.replace(' ', '_')}_"
        f"{dt_desde.strftime('%Y%m%d')}_"
        f"{dt_hasta.strftime('%Y%m%d')}.xlsx"
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )

@router.post("/reparar-fotos/{ficha_id}", response_model=dict)
async def reparar_fotos_ficha(
    ficha_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Toma las fotos HEIC de una ficha específica, las convierte a JPEG,
    las sube a S3 y actualiza las URLs en MongoDB.
    """
    if current_user["rol"] not in ["admin_sede", "super_admin"]:
        raise HTTPException(status_code=403, detail="No tienes permisos")

    # ── Buscar ficha ─────────────────────────────────────────────────────────
    try:
        ficha = await collection_card.find_one({"_id": ObjectId(ficha_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="ID de ficha no válido")

    if not ficha:
        raise HTTPException(status_code=404, detail="Ficha no encontrada")

    fotos     = ficha.get("fotos", {}) or {}
    resultado = {"antes": [], "despues": [], "convertidas": 0, "saltadas": 0, "errores": 0}

    # ── Procesar cada sección ─────────────────────────────────────────────────
    for seccion in ("antes", "despues"):
        urls = fotos.get(seccion, []) or []
        nuevas_urls = []

        for url in urls:
            if not url or not url.startswith("http"):
                nuevas_urls.append(url)
                continue

            # Si ya es JPEG, no tocar
            ext = url.rsplit(".", 1)[-1].lower() if "." in url else ""
            if ext in ("jpg", "jpeg"):
                nuevas_urls.append(url)
                resultado["saltadas"] += 1
                continue

            print(f"  🔄 Procesando foto HEIC: {url}")

            # Descargar
            buf = await descargar_imagen(url)
            if not buf:
                print(f"  ⚠️ No se pudo descargar: {url}")
                nuevas_urls.append(url)
                resultado["errores"] += 1
                continue

            # Convertir con la misma función que ya tienes
            jpeg_buf = comprimir_imagen_para_pdf(buf, max_px=2000, quality=85)
            if not jpeg_buf:
                print(f"  ⚠️ No se pudo convertir: {url}")
                nuevas_urls.append(url)
                resultado["errores"] += 1
                continue

            # Construir nueva key en S3
            from urllib.parse import urlparse
            path     = urlparse(url).path.lstrip("/")
            base_key = path.rsplit(".", 1)[0] if "." in path else path
            nueva_key = f"{base_key}.jpg"

            # Subir a S3
            try:
                jpeg_buf.seek(0)
                s3_client = boto3.client(
                    "s3",
                    region_name=os.getenv("AWS_REGION", "us-east-2"),
                    aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
                    aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
                )
                bucket = os.getenv("AWS_BUCKET_NAME", "rfichas")
                s3_client.put_object(
                    Bucket=bucket,
                    Key=nueva_key,
                    Body=jpeg_buf.read(),
                    ContentType="image/jpeg",
                )
                nueva_url = f"https://{bucket}.s3.{os.getenv('AWS_REGION', 'us-east-2')}.amazonaws.com/{nueva_key}"
                print(f"  ✅ Subida → {nueva_url}")
                nuevas_urls.append(nueva_url)
                resultado["convertidas"] += 1

            except Exception as e:
                print(f"  ❌ Error subiendo a S3: {e}")
                nuevas_urls.append(url)  # mantener original si falla S3
                resultado["errores"] += 1

        resultado[seccion] = nuevas_urls

    # ── Actualizar MongoDB solo si hubo conversiones ──────────────────────────
    if resultado["convertidas"] > 0:
        await collection_card.update_one(
            {"_id": ObjectId(ficha_id)},
            {"$set": {
                "fotos.antes":   resultado["antes"],
                "fotos.despues": resultado["despues"],
            }}
        )
        print(f"  💾 Ficha {ficha_id} actualizada en MongoDB")

    return {
        "success":     True,
        "ficha_id":    ficha_id,
        "convertidas": resultado["convertidas"],
        "saltadas":    resultado["saltadas"],
        "errores":     resultado["errores"],
        "fotos_actualizadas": {
            "antes":   resultado["antes"],
            "despues": resultado["despues"],
        },
        "mensaje": (
            f"✅ {resultado['convertidas']} fotos convertidas a JPEG"
            if resultado["convertidas"] > 0
            else "ℹ️ No había fotos HEIC para convertir"
        )
    }

"""
Endpoint: Reporte de Citas por Sede
GET /reportes/citas/excel?sede_id=SD-40203&fecha_inicio=2026-03-01&fecha_fin=2026-03-31
"""

from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import StreamingResponse
from datetime import date, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Conexión real a MongoDB ────────────────────────────────────────────────────
from motor.motor_asyncio import AsyncIOMotorClient
import os
from dotenv import load_dotenv

load_dotenv()

uri = os.getenv("MONGODB_URI")
db_name = os.getenv("MONGODB_NAME", "DataAgenda")

if not uri:
    raise RuntimeError("MONGODB_URI no está definida en .env")

client = AsyncIOMotorClient(uri)
db = client[db_name]
collection_citas = db["appointments"]
# ──────────────────────────────────────────────────────────────────────────────


COLOR_HEADER    = "1A1A2E"
COLOR_SUBHEADER = "16213E"
COLOR_ACCENT    = "0F3460"
COLOR_ALT_ROW   = "F0F4FF"
COLOR_BORDER    = "CCCCCC"

ESTADO_COLORES = {
    "confirmada": ("D4EDDA", "155724"),
    "completada": ("CCE5FF", "004085"),
    "cancelada":  ("F8D7DA", "721C24"),
    "pendiente":  ("FFF3CD", "856404"),
    "no_asistio": ("E2E3E5", "383D41"),
}
ESTADO_PAGO_COLORES = {
    "pagado":    ("D4EDDA", "155724"),
    "abonado":   ("FFF3CD", "856404"),
    "sin_pago":  ("F8D7DA", "721C24"),
    "pendiente": ("F8D7DA", "721C24"),
}

def thin_border():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def hc(ws, row, col, value, bg=COLOR_HEADER, fg="FFFFFF", bold=True, size=10, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin_border()
    return cell

def dc(ws, row, col, value, bold=False, center=False, bg=None, fg="000000", fmt=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=9)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt
    cell.border = thin_border()
    return cell

def estado_c(ws, row, col, estado, catalog):
    bg, fg = catalog.get(estado, ("FFFFFF", "000000"))
    return dc(ws, row, col, estado.replace("_", " ").upper(), bold=True, center=True, bg=bg, fg=fg)

def cop(v):
    return int(v) if v else 0

def parse_fecha_pago(fp):
    """
    El campo fecha dentro de historial_pagos puede venir como:
      - dict:     {"$date": "2026-03-07T14:28:50.869Z"}
      - string:   "2026-03-07T14:28:50.869Z"
      - datetime: Motor lo deserializa automáticamente en algunos casos
    Devuelve solo YYYY-MM-DD para mostrarlo limpio en el Excel.
    """
    if isinstance(fp, dict):
        fp = fp.get("$date", "")
    if isinstance(fp, datetime):
        return fp.strftime("%Y-%m-%d")
    if isinstance(fp, str) and fp:
        return fp[:10]
    return ""


# ── Constructores de hojas ─────────────────────────────────────────────────────

def build_hoja_citas(ws, citas, sede_nombre, f_ini, f_fin):
    ws.title = "Citas"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = f"REPORTE DE CITAS — {sede_nombre.upper()}"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:P2")
    s = ws["A2"]
    s.value = f"Período: {f_ini}  →  {f_fin}    |    Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    s.font = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    s.fill = PatternFill("solid", fgColor=COLOR_ACCENT)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 4

    COLS = [
        ("#",              5),  ("Fecha",       11), ("Inicio",     9),  ("Fin",        9),
        ("Cliente",       26),  ("Teléfono",    14), ("Email",     28),  ("Profesional",16),
        ("Servicio",      32),  ("Dur. (min)",  11), ("Estado",    13),  ("Abono",      13),
        ("Total",         13),  ("Saldo",       13), ("Pago",      13),  ("Notas",      30),
    ]
    for i, (label, width) in enumerate(COLS, 1):
        hc(ws, 4, i, label, bg=COLOR_SUBHEADER, size=9, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 30

    for i, c in enumerate(citas, 1):
        r = i + 4
        bg = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r].height = 18
        dc(ws, r, 1,  i,                              center=True, bg=bg)
        dc(ws, r, 2,  c.get("fecha"),                 center=True, bg=bg)
        dc(ws, r, 3,  c.get("hora_inicio"),           center=True, bg=bg)
        dc(ws, r, 4,  c.get("hora_fin"),              center=True, bg=bg)
        dc(ws, r, 5,  c.get("cliente_nombre"),        bg=bg)
        dc(ws, r, 6,  c.get("cliente_telefono"),      center=True, bg=bg)
        dc(ws, r, 7,  c.get("cliente_email"),         bg=bg)
        dc(ws, r, 8,  c.get("profesional_nombre"),    bg=bg)
        dc(ws, r, 9,  c.get("servicio_nombre"),       bg=bg, wrap=True)
        dc(ws, r, 10, c.get("servicio_duracion"),     center=True, bg=bg)
        estado_c(ws, r, 11, c.get("estado", ""),      ESTADO_COLORES)
        dc(ws, r, 12, cop(c.get("abono")),            center=True, bg=bg, fmt='#,##0')
        dc(ws, r, 13, cop(c.get("valor_total")),      center=True, bg=bg, fmt='#,##0')
        dc(ws, r, 14, cop(c.get("saldo_pendiente")),  center=True, bg=bg, fmt='#,##0')
        estado_c(ws, r, 15, c.get("estado_pago", ""), ESTADO_PAGO_COLORES)
        dc(ws, r, 16, c.get("notas", ""),             bg=bg, wrap=True)

    ws.freeze_panes = "A5"


def build_hoja_detalle(ws, citas):
    ws.title = "Servicios y Productos"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "DETALLE DE SERVICIOS Y PRODUCTOS POR CITA"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 4

    COLS = [
        ("Fecha", 11), ("Hora", 9), ("Cliente", 26), ("Profesional", 16),
        ("Tipo", 11), ("Ítem", 36), ("Cant.", 7),
        ("Precio Unit.", 14), ("Subtotal", 14), ("Estado Cita", 13), ("# Comprobante", 16),
    ]
    for i, (label, width) in enumerate(COLS, 1):
        hc(ws, 3, i, label, bg=COLOR_SUBHEADER, size=9, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[3].height = 28

    row = 4
    alt = False
    for c in citas:
        items = []
        for sv in c.get("servicios", []):
            items.append(("SERVICIO", sv.get("nombre"), sv.get("cantidad", 1),
                          sv.get("precio", 0), sv.get("subtotal", 0)))
        for pr in c.get("productos", []):
            items.append(("PRODUCTO", pr.get("nombre"), pr.get("cantidad", 1),
                          pr.get("precio_unitario", 0), pr.get("subtotal", 0)))
        bg = COLOR_ALT_ROW if alt else "FFFFFF"
        alt = not alt
        for tipo, nombre, cant, precio, subtotal in items:
            ws.row_dimensions[row].height = 16
            bg_t = "E8F4FD" if tipo == "SERVICIO" else "FFF8E1"
            fg_t = "0D47A1" if tipo == "SERVICIO" else "E65100"
            dc(ws, row, 1,  c.get("fecha"),              center=True, bg=bg)
            dc(ws, row, 2,  c.get("hora_inicio"),        center=True, bg=bg)
            dc(ws, row, 3,  c.get("cliente_nombre"),     bg=bg)
            dc(ws, row, 4,  c.get("profesional_nombre"), bg=bg)
            dc(ws, row, 5,  tipo, center=True, bg=bg_t, fg=fg_t, bold=True)
            dc(ws, row, 6,  nombre, bg=bg, wrap=True)
            dc(ws, row, 7,  cant,     center=True, bg=bg)
            dc(ws, row, 8,  precio,   center=True, bg=bg, fmt='#,##0')
            dc(ws, row, 9,  subtotal, center=True, bg=bg, fmt='#,##0')
            estado_c(ws, row, 10, c.get("estado", ""), ESTADO_COLORES)
            dc(ws, row, 11, c.get("numero_comprobante") or "—", center=True, bg=bg)
            row += 1
    ws.freeze_panes = "A4"


def build_hoja_pagos(ws, citas):
    ws.title = "Historial de Pagos"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = "HISTORIAL DE PAGOS POR CITA"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 4

    COLS = [
        ("Fecha Cita", 11), ("Hora", 9), ("Cliente", 26), ("Profesional", 16),
        ("Fecha Pago", 18), ("Monto", 13), ("Método", 14),
        ("Tipo", 16), ("Registrado por", 30), ("Saldo Después", 14),
    ]
    for i, (label, width) in enumerate(COLS, 1):
        hc(ws, 3, i, label, bg=COLOR_SUBHEADER, size=9, wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[3].height = 28

    row = 4
    alt = False
    for c in citas:
        pagos = c.get("historial_pagos", [])
        if not pagos:
            continue
        bg = COLOR_ALT_ROW if alt else "FFFFFF"
        alt = not alt
        for p in pagos:
            ws.row_dimensions[row].height = 16
            dc(ws, row, 1,  c.get("fecha"),                     center=True, bg=bg)
            dc(ws, row, 2,  c.get("hora_inicio"),               center=True, bg=bg)
            dc(ws, row, 3,  c.get("cliente_nombre"),            bg=bg)
            dc(ws, row, 4,  c.get("profesional_nombre"),        bg=bg)
            dc(ws, row, 5,  parse_fecha_pago(p.get("fecha")),   center=True, bg=bg)
            dc(ws, row, 6,  cop(p.get("monto")),                center=True, bg=bg, fmt='#,##0')
            dc(ws, row, 7,  p.get("metodo", "").replace("_", " "), center=True, bg=bg)
            dc(ws, row, 8,  p.get("tipo",   "").replace("_", " "), center=True, bg=bg)
            dc(ws, row, 9,  p.get("registrado_por", ""),        bg=bg)
            dc(ws, row, 10, cop(p.get("saldo_despues")),        center=True, bg=bg, fmt='#,##0')
            row += 1
    ws.freeze_panes = "A4"


def generar_excel(citas, sede_nombre, f_ini, f_fin):
    wb = openpyxl.Workbook()
    build_hoja_citas(wb.active, citas, sede_nombre, f_ini, f_fin)
    build_hoja_detalle(wb.create_sheet(), citas)
    build_hoja_pagos(wb.create_sheet(), citas)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Endpoint ───────────────────────────────────────────────────────────────────

@router.get("/reportes/citas/excel")
async def reporte_citas_excel(
    sede_id: str = Query(..., description="ID de la sede, ej: SD-40203"),
    fecha_inicio: date = Query(..., description="YYYY-MM-DD"),
    fecha_fin:    date = Query(..., description="YYYY-MM-DD"),
):
    if fecha_fin < fecha_inicio:
        raise HTTPException(400, "fecha_fin debe ser >= fecha_inicio")

    # ── Query real a MongoDB ───────────────────────────────────────────────────
    # El campo `fecha` en appointments es un string "YYYY-MM-DD",
    # la comparación lexicográfica funciona perfectamente.
    cursor = collection_citas.find(
        {
            "sede_id": sede_id,
            "fecha": {
                "$gte": str(fecha_inicio),
                "$lte": str(fecha_fin),
            },
        },
        {"_id": 0},                           # excluimos _id
    ).sort([("fecha", 1), ("hora_inicio", 1)])

    citas = await cursor.to_list(length=None)  # trae todos los documentos
    # ──────────────────────────────────────────────────────────────────────────

    if not citas:
        raise HTTPException(
            404,
            f"No hay citas para la sede {sede_id} entre {fecha_inicio} y {fecha_fin}",
        )

    sede_nombre = citas[0].get("sede_nombre", sede_id)
    buf = generar_excel(citas, sede_nombre, str(fecha_inicio), str(fecha_fin))
    fname = f"reporte_citas_{sede_id}_{fecha_inicio}_al_{fecha_fin}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

# ── datos mock para pruebas (reemplaza con MongoDB) ───────────────────────────

def _mock_citas():
    return [
        {
            "sede_id": "SD-40203", "sede_nombre": "RF AV ORIENTAL - MEDELLIN",
            "cliente_nombre": "Luz marina Gómez González",
            "cliente_email": "Mg7221780@gmail.com", "cliente_telefono": "3003093055",
            "profesional_nombre": "Diana", "servicio_nombre": "COMPLETO EXTRA ALTA P.A",
            "servicios": [{"nombre": "COMPLETO EXTRA ALTA P.A", "precio": 187000, "cantidad": 1, "subtotal": 187000}],
            "productos": [],
            "fecha": "2026-03-27", "hora_inicio": "10:00", "hora_fin": "12:00",
            "servicio_duracion": 180, "estado": "confirmada", "estado_pago": "abonado",
            "abono": 50000, "valor_total": 187000, "saldo_pendiente": 137000,
            "notas": "Cliente va a reagendar",
            "historial_pagos": [{"fecha": {"$date": "2026-03-07T14:28:50Z"}, "monto": 50000,
                                  "metodo": "transferencia", "tipo": "abono_inicial",
                                  "registrado_por": "rizoscentro@gmail.com", "saldo_despues": 137000}],
            "numero_comprobante": None,
        },
        {
            "sede_id": "SD-40203", "sede_nombre": "RF AV ORIENTAL - MEDELLIN",
            "cliente_nombre": "Valentina Torres Reyes",
            "cliente_email": "valtorres@gmail.com", "cliente_telefono": "3112345678",
            "profesional_nombre": "Diana", "servicio_nombre": "CORTE ESTANDAR P.A",
            "servicios": [{"nombre": "CORTE ESTANDAR P.A", "precio": 97000, "cantidad": 1, "subtotal": 97000}],
            "productos": [{"nombre": "SPECIAL CREMA 3 EN 1 250 ML", "cantidad": 1,
                           "precio_unitario": 78000, "subtotal": 78000}],
            "fecha": "2026-03-27", "hora_inicio": "13:00", "hora_fin": "14:30",
            "servicio_duracion": 90, "estado": "completada", "estado_pago": "pagado",
            "abono": 175000, "valor_total": 175000, "saldo_pendiente": 0, "notas": "",
            "historial_pagos": [{"fecha": {"$date": "2026-03-27T14:30:00Z"}, "monto": 175000,
                                  "metodo": "efectivo", "tipo": "pago_completo",
                                  "registrado_por": "rizoscentro@gmail.com", "saldo_despues": 0}],
            "numero_comprobante": "40732050",
        },
        {
            "sede_id": "SD-40203", "sede_nombre": "RF AV ORIENTAL - MEDELLIN",
            "cliente_nombre": "Camila Ríos Montoya",
            "cliente_email": "crios@hotmail.com", "cliente_telefono": "3209876543",
            "profesional_nombre": "Sandra", "servicio_nombre": "BRUSHING P.A",
            "servicios": [{"nombre": "BRUSHING P.A", "precio": 65000, "cantidad": 1, "subtotal": 65000}],
            "productos": [],
            "fecha": "2026-03-28", "hora_inicio": "09:00", "hora_fin": "10:00",
            "servicio_duracion": 60, "estado": "cancelada", "estado_pago": "sin_pago",
            "abono": 0, "valor_total": 65000, "saldo_pendiente": 65000,
            "notas": "Canceló por enfermedad", "historial_pagos": [],
            "numero_comprobante": None,
        },
    ]