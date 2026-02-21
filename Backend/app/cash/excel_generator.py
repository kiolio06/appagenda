# ============================================================
# excel_generator.py - ACTUALIZADO CON 3 HOJAS ADICIONALES
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import Dict, List

# ============================================================
# FUNCIÓN PRINCIPAL ACTUALIZADA
# ============================================================

def _obtener_periodo(resumen: Dict) -> tuple[str, str]:
    """Resuelve el período del reporte soportando día único o rango."""
    fecha_inicio = str(resumen.get("fecha_inicio") or resumen.get("fecha") or "")
    fecha_fin = str(resumen.get("fecha_fin") or resumen.get("fecha") or fecha_inicio)
    return fecha_inicio, fecha_fin

def generar_reporte_excel_caja_completo(
    resumen: Dict,
    sede_info: Dict,
    facturas: List[Dict],
    egresos: List[Dict],
    movimientos_efectivo: Dict
) -> BytesIO:
    """
    Genera Excel con 4 hojas:
    1. Resumen de Caja (existente)
    2. Resumen Flujo de Ingresos (nueva)
    3. Resumen Flujo de Egresos (nueva)
    4. Movimientos Efectivo (nueva)
    """
    
    wb = Workbook()
    fecha_inicio, fecha_fin = _obtener_periodo(resumen)
    
    # Hoja 1: Resumen de Caja (la que ya teníamos)
    ws_resumen = wb.active
    ws_resumen.title = "Resumen de Caja"
    _crear_hoja_resumen_caja(ws_resumen, resumen, sede_info)
    
    # Hoja 2: Flujo de Ingresos
    ws_ingresos = wb.create_sheet("Flujo de Ingresos")
    _crear_hoja_flujo_ingresos(ws_ingresos, sede_info, fecha_inicio, fecha_fin, facturas)
    
    # Hoja 3: Flujo de Egresos
    ws_egresos = wb.create_sheet("Flujo de Egresos")
    _crear_hoja_flujo_egresos(ws_egresos, sede_info, fecha_inicio, fecha_fin, egresos)
    
    # Hoja 4: Movimientos Efectivo
    ws_movimientos = wb.create_sheet("Movimientos Efectivo")
    _crear_hoja_movimientos_efectivo(ws_movimientos, sede_info, fecha_inicio, fecha_fin, movimientos_efectivo)
    
    # Guardar en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

# ============================================================
# HOJA 1: RESUMEN DE CAJA (YA EXISTENTE)
# ============================================================

def _crear_hoja_resumen_caja(ws, resumen: Dict, sede_info: Dict):
    """Crea la hoja de resumen de caja (código anterior)"""
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True)
    subtitulo_font = Font(name='Arial', size=12, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', size=11, bold=True)
    
    centro = Alignment(horizontal='center', vertical='center')
    derecha = Alignment(horizontal='right', vertical='center')
    
    borde_grueso_abajo = Border(bottom=Side(style='medium'))
    
    relleno_gris = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    relleno_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    relleno_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    fila = 1
    
    # Título
    ws.merge_cells(f'A{fila}:D{fila}')
    ws[f'A{fila}'] = "RESUMEN DE CAJA DE VENTAS"
    ws[f'A{fila}'].font = titulo_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Empresa
    ws.merge_cells(f'A{fila}:D{fila}')
    ws[f'A{fila}'] = sede_info.get("razon_social", "SALÓN RIZOS FELICES CL SAS")
    ws[f'A{fila}'].font = subtitulo_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Dirección
    ws.merge_cells(f'A{fila}:D{fila}')
    direccion_completa = f"{sede_info.get('direccion', '')}, {sede_info.get('ciudad', '')}, {sede_info.get('pais', '')}"
    ws[f'A{fila}'] = direccion_completa
    ws[f'A{fila}'].font = normal_font
    ws[f'A{fila}'].alignment = centro
    fila += 2
    
    # Periodo
    fecha_inicio, fecha_fin = _obtener_periodo(resumen)
    ws[f'A{fila}'] = "Inicio:"
    ws[f'B{fila}'] = f"{fecha_inicio} 00:00"
    ws[f'A{fila}'].font = header_font
    fila += 1
    
    ws[f'A{fila}'] = "Fin:"
    ws[f'B{fila}'] = f"{fecha_fin} 23:59"
    ws[f'A{fila}'].font = header_font
    fila += 2
    
    # Línea
    ws.merge_cells(f'A{fila}:D{fila}')
    ws[f'A{fila}'].border = borde_grueso_abajo
    fila += 1
    
    # Saldo inicial
    ws[f'A{fila}'] = "SALDO INICIAL EN EFECTIVO"
    ws[f'A{fila}'].font = header_font
    ws[f'D{fila}'] = resumen["efectivo_inicial"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 2
    
    # Ingresos
    ws[f'A{fila}'] = "INGRESOS"
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].fill = relleno_gris
    fila += 1
    
    ws[f'A{fila}'] = "- Efectivo"
    ws[f'D{fila}'] = resumen["ingresos_efectivo"]["total"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'A{fila}'] = "- Abonos a Reservas"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["abonos"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1

    ws[f'A{fila}'] = "- Tarjeta Crédito"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["tarjeta_credito"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'A{fila}'] = "- Tarjeta Débito"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["tarjeta_debito"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1

    ws[f'A{fila}'] = "- POS"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["pos"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1

    ws[f'A{fila}'] = "- Link de Pago"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["link_de_pago"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1

    ws[f'A{fila}'] = "- Giftcard"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["giftcard"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1

    ws[f'A{fila}'] = "- Addi"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["addi"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'A{fila}'] = "- Transferencias"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["transferencia"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'A{fila}'] = "- Otros"
    ws[f'D{fila}'] = resumen["ingresos_otros_metodos"]["otros"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'D{fila}'].border = Border(top=Side(style='thin'))
    fila += 1
    
    ws[f'A{fila}'] = "Total Ingresos (+)"
    ws[f'A{fila}'].font = total_font
    ws[f'D{fila}'] = resumen["total_vendido"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    ws[f'D{fila}'].font = total_font
    fila += 2
    
    # Egresos
    ws[f'A{fila}'] = "EGRESOS"
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].fill = relleno_gris
    fila += 1
    
    ws[f'A{fila}'] = "- Compras Internas"
    ws[f'D{fila}'] = resumen["egresos"]["compras_internas"]["total"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'A{fila}'] = "- Gastos Operativos"
    ws[f'D{fila}'] = resumen["egresos"]["gastos_operativos"]["total"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'A{fila}'] = "- Retiros"
    ws[f'D{fila}'] = resumen["egresos"]["retiros_caja"]["total"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    fila += 1
    
    ws[f'D{fila}'].border = Border(top=Side(style='thin'))
    fila += 1
    
    ws[f'A{fila}'] = "Total Egresos (-)"
    ws[f'A{fila}'].font = total_font
    ws[f'D{fila}'] = resumen["egresos"]["total"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    ws[f'D{fila}'].font = total_font
    fila += 2
    
    # Línea
    ws.merge_cells(f'A{fila}:D{fila}')
    ws[f'A{fila}'].border = borde_grueso_abajo
    fila += 1
    
    # Resultado
    resultado = resumen["total_vendido"] - resumen["egresos"]["total"]
    
    ws[f'A{fila}'] = "RESULTADO DEL PERÍODO (=)"
    ws[f'A{fila}'].font = total_font
    ws[f'D{fila}'] = resultado
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    ws[f'D{fila}'].font = total_font
    ws[f'D{fila}'].fill = relleno_verde if resultado >= 0 else relleno_rojo
    fila += 2
    
    # Saldo final
    ws[f'A{fila}'] = "SALDO FINAL EN EFECTIVO"
    ws[f'A{fila}'].font = header_font
    ws[f'D{fila}'] = resumen["efectivo_esperado"]
    ws[f'D{fila}'].number_format = '#,##0.00'
    ws[f'D{fila}'].alignment = derecha
    ws[f'D{fila}'].font = total_font
    ws[f'D{fila}'].fill = relleno_verde if resumen["efectivo_esperado"] >= 0 else relleno_rojo
    
    # Anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

# ============================================================
# HOJA 2: FLUJO DE INGRESOS
# ============================================================

def _crear_hoja_flujo_ingresos(
    ws,
    sede_info: Dict,
    fecha_inicio: str,
    fecha_fin: str,
    facturas: List[Dict]
):
    """Crea la hoja de flujo de ingresos"""
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=9)
    centro = Alignment(horizontal='center', vertical='center')
    
    fila = 1
    
    # Título
    ws.merge_cells(f'A{fila}:N{fila}')
    ws[f'A{fila}'] = "Resumen Flujo de Ingresos"
    ws[f'A{fila}'].font = titulo_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Empresa
    ws.merge_cells(f'A{fila}:N{fila}')
    ws[f'A{fila}'] = sede_info.get("razon_social", "SALÓN RIZOS FELICES CL SAS")
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Dirección
    ws.merge_cells(f'A{fila}:N{fila}')
    direccion = f"{sede_info.get('direccion', '')}, {sede_info.get('ciudad', '')}, {sede_info.get('pais', '')}"
    ws[f'A{fila}'] = direccion
    ws[f'A{fila}'].alignment = centro
    fila += 2
    
    # Periodo
    ws[f'A{fila}'] = "Inicio"
    ws[f'B{fila}'] = f"{fecha_inicio} 00:00"
    fila += 1
    ws[f'A{fila}'] = "Fin"
    ws[f'B{fila}'] = f"{fecha_fin} 23:59"
    fila += 2
    
    # Headers
    headers = [
        "Fecha", "Nombre cliente", "C.I. cliente", "Email cliente", "Teléfono cliente",
        "Medio de Pago", "Tipo de Movimiento", "ID Movimiento",
        "Nro Comprobante", "Flujo del periodo",
        "Usuario última modificación"
    ]
    
    for col, header in enumerate(headers, start=1):
        celda = ws.cell(row=fila, column=col)
        celda.value = header
        celda.font = header_font
        celda.alignment = centro
    
    fila += 1
    
    # Datos
    for factura in facturas:
        fecha_hora = factura["fecha"].strftime("%d/%m/%Y %H:%M") if factura["fecha"] else ""
        
        ws.cell(row=fila, column=1).value = fecha_hora
        ws.cell(row=fila, column=2).value = factura["nombre_cliente"]
        ws.cell(row=fila, column=3).value = factura["cedula_cliente"]
        ws.cell(row=fila, column=4).value = factura["email_cliente"]
        ws.cell(row=fila, column=5).value = factura["telefono_cliente"]
        ws.cell(row=fila, column=6).value = factura["medio_pago"]
        ws.cell(row=fila, column=7).value = factura["tipo_movimiento"]
        ws.cell(row=fila, column=8).value = factura["id_movimiento"]
        ws.cell(row=fila, column=9).value = factura["nro_comprobante"]
        ws.cell(row=fila, column=10).value = factura["flujo_periodo"]
        ws.cell(row=fila, column=10).number_format = '#,##0'
        ws.cell(row=fila, column=11).value = factura["usuario_modificacion"]
        
        fila += 1
    
    # Anchos de columna
    anchos = [18, 30, 15, 30, 15, 25, 20, 15, 15, 15, 15, 30, 25, 20]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

# ============================================================
# HOJA 3: FLUJO DE EGRESOS
# ============================================================

def _crear_hoja_flujo_egresos(
    ws,
    sede_info: Dict,
    fecha_inicio: str,
    fecha_fin: str,
    egresos: List[Dict]
):
    """Crea la hoja de flujo de egresos"""
    
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    centro = Alignment(horizontal='center', vertical='center')
    
    fila = 1
    
    # Título
    ws.merge_cells(f'A{fila}:H{fila}')
    ws[f'A{fila}'] = "Resumen Flujo de Egresos"
    ws[f'A{fila}'].font = titulo_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Empresa
    ws.merge_cells(f'A{fila}:H{fila}')
    ws[f'A{fila}'] = sede_info.get("razon_social", "SALÓN RIZOS FELICES CL SAS")
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Dirección
    ws.merge_cells(f'A{fila}:H{fila}')
    direccion = f"{sede_info.get('direccion', '')}, {sede_info.get('ciudad', '')}, {sede_info.get('pais', '')}"
    ws[f'A{fila}'] = direccion
    ws[f'A{fila}'].alignment = centro
    fila += 2
    
    # Periodo
    ws[f'A{fila}'] = "Inicio"
    ws[f'B{fila}'] = f"{fecha_inicio} 00:00"
    fila += 1
    ws[f'A{fila}'] = "Fin"
    ws[f'B{fila}'] = f"{fecha_fin} 23:59"
    fila += 2
    
    # Headers
    headers = [
        "Fecha", "Concepto", "Medio de Pago", "Tipo de Movimiento", "ID Egreso",
        "Nro Comprobante", "Flujo del periodo (-)", "Notas"
    ]
    
    for col, header in enumerate(headers, start=1):
        celda = ws.cell(row=fila, column=col)
        celda.value = header
        celda.font = header_font
        celda.alignment = centro
    
    fila += 1
    
    # Datos
    for egreso in egresos:
        fecha_hora = egreso["fecha"].strftime("%d/%m/%Y %H:%M") if egreso["fecha"] else ""
        
        ws.cell(row=fila, column=1).value = fecha_hora
        ws.cell(row=fila, column=2).value= egreso["concepto"]
        ws.cell(row=fila, column=3).value = egreso["medio_pago"]
        ws.cell(row=fila, column=4).value = egreso["tipo_movimiento"]
        ws.cell(row=fila, column=5).value = egreso["id_egreso"]
        ws.cell(row=fila, column=6).value = egreso["nro_comprobante"]
        ws.cell(row=fila, column=7).value = egreso["flujo_periodo"]
        ws.cell(row=fila, column=7).number_format = '#,##0'
        ws.cell(row=fila, column=8).value = egreso["notas"]
        
        fila += 1
    
    # Anchos
    anchos = [18, 20, 20, 20, 25, 20, 18, 50]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

# ============================================================
# HOJA 4: MOVIMIENTOS EFECTIVO
# ============================================================

def _crear_hoja_movimientos_efectivo(
    ws,
    sede_info: Dict,
    fecha_inicio: str,
    fecha_fin: str,
    movimientos: Dict
):
    """Crea la hoja de movimientos en efectivo con saldo corrido"""
    
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    total_font = Font(name='Arial', size=11, bold=True)
    centro = Alignment(horizontal='center', vertical='center')
    derecha = Alignment(horizontal='right', vertical='center')
    
    relleno_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    fila = 1
    
    # Título
    ws.merge_cells(f'A{fila}:G{fila}')
    ws[f'A{fila}'] = "Movimientos en Efectivo"
    ws[f'A{fila}'].font = titulo_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Empresa
    ws.merge_cells(f'A{fila}:G{fila}')
    ws[f'A{fila}'] = sede_info.get("razon_social", "SALÓN RIZOS FELICES CL SAS")
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].alignment = centro
    fila += 1
    
    # Dirección
    ws.merge_cells(f'A{fila}:G{fila}')
    direccion = f"{sede_info.get('direccion', '')}, {sede_info.get('ciudad', '')}, {sede_info.get('pais', '')}"
    ws[f'A{fila}'] = direccion
    ws[f'A{fila}'].alignment = centro
    fila += 2
    
    # Periodo
    ws[f'A{fila}'] = "Inicio"
    ws[f'B{fila}'] = f"{fecha_inicio} 00:00"
    fila += 1
    ws[f'A{fila}'] = "Fin"
    ws[f'B{fila}'] = f"{fecha_fin} 23:59"
    fila += 2
    
    # Saldo inicial
    ws[f'A{fila}'] = "SALDO INICIAL"
    ws[f'A{fila}'].font = total_font
    ws[f'G{fila}'] = movimientos["saldo_inicial"]
    ws[f'G{fila}'].number_format = '#,##0.00'
    ws[f'G{fila}'].alignment = derecha
    ws[f'G{fila}'].font = total_font
    ws[f'G{fila}'].fill = relleno_verde
    fila += 2
    
    # Headers
    headers = [
        "Fecha", "Tipo", "Descripción", "Comprobante",
        "Ingreso (+)", "Egreso (-)", "Saldo"
    ]
    
    for col, header in enumerate(headers, start=1):
        celda = ws.cell(row=fila, column=col)
        celda.value = header
        celda.font = header_font
        celda.alignment = centro
    
    fila += 1
    
    # Movimientos
    for mov in movimientos["movimientos"]:
        fecha_hora = mov["fecha"].strftime("%d/%m/%Y %H:%M") if mov["fecha"] else ""
        
        ws.cell(row=fila, column=1).value = fecha_hora
        ws.cell(row=fila, column=2).value = mov["tipo"]
        ws.cell(row=fila, column=3).value = mov["descripcion"]
        ws.cell(row=fila, column=4).value = mov["comprobante"]
        ws.cell(row=fila, column=5).value = mov["ingreso"]
        ws.cell(row=fila, column=5).number_format = '#,##0.00'
        ws.cell(row=fila, column=6).value = mov["egreso"]
        ws.cell(row=fila, column=6).number_format = '#,##0.00'
        ws.cell(row=fila, column=7).value = mov["saldo"]
        ws.cell(row=fila, column=7).number_format = '#,##0.00'
        ws.cell(row=fila, column=7).alignment = derecha
        
        fila += 1
    
    fila += 1
    
    # Saldo final
    ws[f'A{fila}'] = "SALDO FINAL"
    ws[f'A{fila}'].font = total_font
    ws[f'G{fila}'] = movimientos["saldo_final"]
    ws[f'G{fila}'].number_format = '#,##0.00'
    ws[f'G{fila}'].alignment = derecha
    ws[f'G{fila}'].font = total_font
    ws[f'G{fila}'].fill = relleno_verde
    
    # Anchos
    anchos = [18, 15, 40, 15, 15, 15, 15]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

# ============================================================
# FUNCIÓN HELPER PARA NOMBRES DE ARCHIVO
# ============================================================

def generar_nombre_archivo_excel(nombre: str, fecha: str) -> str:
    """Genera nombre de archivo Excel descriptivo"""
    return f"Reporte_Caja_{nombre}_{fecha}.xlsx"
